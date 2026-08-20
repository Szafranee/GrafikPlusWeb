import tempfile
import unittest
from pathlib import Path
from unittest.mock import Mock

import requests
from openpyxl import load_workbook

from backend.config import DEFAULT_INSTALLATION_FILENAME, ScheduleConfig
from backend.schedule_scraper import LoginError, ScheduleFetchError, ScheduleScraper


FIXTURES = Path(__file__).parent / "fixtures"
LOGIN_URL = "https://gpt.canalplus.pl/Account/Login"
PERSONAL_URL = (
    "https://gpt.canalplus.pl/User/Schedule"
    "?date=05%2F18%2F2026%2000%3A00%3A00"
)
GENERAL_URL = (
    "https://gpt.canalplus.pl/Schedule/Editing"
    "?date=05%2F18%2F2026%2000%3A00%3A00"
)


def _response(text: str) -> Mock:
    response = Mock()
    response.text = text
    response.raise_for_status.return_value = None
    return response


class ScheduleScraperTests(unittest.TestCase):
    def _config(self, directory: str, is_personal: bool) -> ScheduleConfig:
        return ScheduleConfig(
            username="jan.kowalski",
            password="secret",
            output_dir=directory,
            output_filename=DEFAULT_INSTALLATION_FILENAME,
            start_date="2026-05-20",
            end_date="2026-05-20",
            is_personal=is_personal,
            use_template_export=False,
        )

    def _run_successful_scrape(self, is_personal: bool, fixture: str, url: str):
        with tempfile.TemporaryDirectory() as directory:
            config = self._config(directory, is_personal)
            scraper = ScheduleScraper(config)
            scraper.session = Mock()
            scraper.session.post.return_value = _response("Zalogowano")
            scraper.session.get.return_value = _response(
                (FIXTURES / fixture).read_text(encoding="utf-8")
            )

            filename = scraper.scrape_schedule()

            self.assertEqual(filename, DEFAULT_INSTALLATION_FILENAME)
            scraper.session.post.assert_called_once_with(
                LOGIN_URL,
                data={"username": "jan.kowalski", "password": "secret"},
                timeout=30,
            )
            scraper.session.get.assert_called_once_with(url, timeout=30)

            output_path = Path(directory) / DEFAULT_INSTALLATION_FILENAME
            self.assertTrue(output_path.is_file())
            workbook = load_workbook(output_path, read_only=True)
            try:
                rows = list(workbook.active.iter_rows(values_only=True))
            finally:
                workbook.close()
            self.assertEqual(len(rows), 2)

    def test_personal_endpoint_login_parsing_and_export(self):
        self._run_successful_scrape(
            is_personal=True,
            fixture="personal_schedule.html",
            url=PERSONAL_URL,
        )

    def test_general_endpoint_login_parsing_and_export(self):
        self._run_successful_scrape(
            is_personal=False,
            fixture="general_schedule.html",
            url=GENERAL_URL,
        )

    def test_invalid_credentials_stop_before_schedule_request(self):
        with tempfile.TemporaryDirectory() as directory:
            scraper = ScheduleScraper(self._config(directory, is_personal=True))
            scraper.session = Mock()
            scraper.session.post.return_value = _response(
                "Niepoprawny identyfikator lub hasło."
            )

            with self.assertRaises(LoginError):
                scraper.scrape_schedule()

            scraper.session.get.assert_not_called()

    def test_schedule_http_error_becomes_fetch_error(self):
        with tempfile.TemporaryDirectory() as directory:
            scraper = ScheduleScraper(self._config(directory, is_personal=True))
            scraper.session = Mock()
            scraper.session.post.return_value = _response("Zalogowano")
            failed_response = _response("")
            failed_response.raise_for_status.side_effect = requests.HTTPError("503")
            scraper.session.get.return_value = failed_response

            with self.assertRaises(ScheduleFetchError):
                scraper.scrape_schedule()

    def test_unrecognized_html_contract_becomes_fetch_error(self):
        with tempfile.TemporaryDirectory() as directory:
            scraper = ScheduleScraper(self._config(directory, is_personal=True))
            scraper.session = Mock()
            scraper.session.post.return_value = _response("Zalogowano")
            scraper.session.get.return_value = _response(
                "<html><body><p>Struktura strony uległa zmianie</p></body></html>"
            )

            with self.assertRaises(ScheduleFetchError):
                scraper.scrape_schedule()

            self.assertFalse(
                (Path(directory) / DEFAULT_INSTALLATION_FILENAME).exists()
            )


if __name__ == "__main__":
    unittest.main()
