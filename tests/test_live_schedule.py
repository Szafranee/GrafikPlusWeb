import os
import tempfile
import unittest
from datetime import date
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook

from backend.config import DEFAULT_INSTALLATION_FILENAME, ScheduleConfig
from backend.schedule_scraper import ScheduleScraper


PROJECT_ROOT = Path(__file__).resolve().parents[1]
load_dotenv(PROJECT_ROOT / ".env")


def _enabled(value: str | None) -> bool:
    return (value or "").strip().lower() in {"1", "true", "yes", "on"}


@unittest.skipUnless(
    _enabled(os.environ.get("RUN_LIVE_SCHEDULE_TESTS")),
    "Set RUN_LIVE_SCHEDULE_TESTS=true to test the current external service.",
)
class LiveScheduleTests(unittest.TestCase):
    def test_current_service_login_url_structure_parsing_and_export(self):
        username = os.environ.get("LIVE_SCHEDULE_USERNAME", "").strip()
        password = os.environ.get("LIVE_SCHEDULE_PASSWORD", "")
        if not username or not password:
            self.skipTest("LIVE_SCHEDULE_USERNAME and LIVE_SCHEDULE_PASSWORD are required.")

        mode = (
            os.environ.get("LIVE_SCHEDULE_MODE", "").strip().lower() or "personal"
        )
        if mode not in {"personal", "installation"}:
            self.fail("LIVE_SCHEDULE_MODE must be 'personal' or 'installation'.")
        schedule_date = (
            os.environ.get("LIVE_SCHEDULE_DATE", "").strip()
            or date.today().isoformat()
        )

        with tempfile.TemporaryDirectory() as directory:
            config = ScheduleConfig(
                username=username,
                password=password,
                output_dir=directory,
                output_filename=DEFAULT_INSTALLATION_FILENAME,
                start_date=schedule_date,
                end_date=schedule_date,
                is_personal=mode == "personal",
                use_template_export=False,
            )

            filename = ScheduleScraper(config).scrape_schedule()

            self.assertEqual(filename, DEFAULT_INSTALLATION_FILENAME)
            workbook = load_workbook(
                Path(directory) / DEFAULT_INSTALLATION_FILENAME,
                read_only=True,
            )
            try:
                row_count = sum(1 for _ in workbook.active.iter_rows(values_only=True))
            finally:
                workbook.close()
            self.assertGreater(
                row_count,
                1,
                "The live page returned no parseable schedule rows; its HTML contract may have changed.",
            )


if __name__ == "__main__":
    unittest.main()
